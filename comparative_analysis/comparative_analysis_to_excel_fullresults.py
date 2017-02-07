# excel and excel from pd df
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

from sqlalchemy import create_engine
import pandas as pd


# base_id: [lookup_table, base_name]
BASE_ID_LOOKUP = {
    'site_ref': 'pl._lookup_cpd',
    'realm_biome': 'pl._lookup_realm_biome',
    'sitrecid': 'pl._lookup_kba_iba',
    'aze_id': 'pl._lookup_aze'
    }


# base: [base_table, base_id, lookup_table, lookup_id, lookup_name]
BASE_LOOKUP = {'eba': ['eba','ebaname', 'eba', 'ebaname', 'ebaname'],
      'feow': ['feow', 'ecoregion', 'feow', 'ecoregion', 'ecoregion'],
      'g200_terr': ['g200_terr', 'g200_regio', 'g200_terr', 'g200_regio', 'g200_regio'],
      'g200_marine': ['g200_marine', 'g200_regio', 'g200_marine', 'g200_regio', 'g200_regio'],
      'g200_fw': ['g200_fw', 'g200_regio', 'g200_fw', 'g200_regio', 'g200_regio'],
      'meow_200m_p': ['meow_200m', 'province', 'meow_200m', 'province', 'province'],
      'meow_200m_e': ['meow_200m', 'ecoregion', 'meow_200m', 'ecoregion', 'ecoregion'],
      'cpd': ['cpd', 'site_ref', '_lookup_cpd', 'id', 'name'],
      'udv_dcw_p': ['udv_dcw', 'provname', 'udv_dcw', 'provname', 'provname'],
      'udv_dcw_b': ['udv_dcw', 'biome', '_lookup_udv_biomename', 'biome', 'biomename'],
      'udv_dcw_r': ['udv_dcw', 'realm', '_lookup_wwf_udv_realm', 'realm', 'realm_name'],
      'wwf_e': ['wwf_terr_ecos_clean', 'eco_name', 'wwf_terr_ecos_clean', 'eco_name', 'eco_name'],
      'wwf_b': ['wwf_terr_ecos_clean', 'biome', '_lookup_wwf_terr_biome', 'bioid', 'name'],
      'wwf_r': ['wwf_terr_ecos_clean', 'realm', '_lookup_wwf_udv_realm', 'realm', 'realm_name'],
      'wwf_rb': ['wwf_terr_ecos_clean', 'realm_biome', '_lookup_realm_biome', 'id', 'name'],
      'hs': ['hotspot_type_core', 'fname', 'hotspot_type_core', 'fname', 'fname'],
      'ws': ['wilderness_type_core', 'wa_name', 'wilderness_type_core', 'wa_name', 'wa_name'],
      'aze': ['aze', 'aze_id', '_lookup_aze', 'id', 'name'],
      'iba': ['iba', 'sitrecid', '_lookup_kba_iba', 'id', 'name'],
      'kba': ['kba', 'sitrecid', '_lookup_kba_iba', 'id', 'name']}

# start
engine = create_engine('postgresql://postgres:gisintern@localhost/whs_v2')

def gen_sql_tls_percentage_overlap(base):
    # join tables: wh attrs and pl lookups
    if base not in BASE_LOOKUP:
        raise Exception('base not in the base_lookup')

    base_table = BASE_LOOKUP[base][0]
    base_id = BASE_LOOKUP[base][1]
    lookup_table = BASE_LOOKUP[base][2]
    lookup_id = BASE_LOOKUP[base][3]
    lookup_name = BASE_LOOKUP[base][4]

    # arguments for constructing sql
    tls_lookup_table = TLS_SHAPE
    intersect = TLS_SCHEMA + '.intersect_' + base_table + '_' + base_id

    pl_lookup_table = 'pl.' + lookup_table
    pl_lookup_table_id = lookup_id
    pl_lookup_table_name = lookup_name

    # sql
    sql = """
    SELECT DISTINCT {1}.id, {1}.name as tsl_name,
    {2}.{4}

    FROM
    {0},{1},{2}

    WHERE
    {0}.tid = {1}.id AND
    {0}.bid = {2}.{3}

    """.format(intersect, tls_lookup_table, 
        pl_lookup_table, pl_lookup_table_id, pl_lookup_table_name)

    return sql

def tls_get_base_df(base):
    sql = gen_sql_tls_percentage_overlap(base)
    return pd.read_sql(sql, engine)

def gen_sql_wh_percentage_overlap(base):
    # join tables: wh attrs and pl lookups
    if base not in BASE_LOOKUP:
        raise Exception('base not in the base_lookup')

    base_table = BASE_LOOKUP[base][0]
    base_id = BASE_LOOKUP[base][1]
    lookup_table = BASE_LOOKUP[base][2]
    lookup_id = BASE_LOOKUP[base][3]
    lookup_name = BASE_LOOKUP[base][4]

    # arguments for constructing sql
    wh_lookup_table = OUTPUT_SCHEMA + '.' + COMBINED_WH_NOMINATION_VIEW
    intersect = OUTPUT_SCHEMA + '.intersect_' + base_table + '_' + base_id

    pl_lookup_table = 'pl.' + lookup_table
    pl_lookup_table_id = lookup_id
    pl_lookup_table_name = lookup_name

    # sql
    sql = """
    SELECT DISTINCT wdpaid, en_name,
    {2}.{4},
    seg_area as overlap_area,
    theme_area as wh_area,
    base_area,
    seg_area/theme_area as per_wh

    FROM
    {0},{1},{2}

    WHERE
    {0}.tid = {1}.wdpaid AND
    {0}.bid = {2}.{3}

    """.format(intersect, wh_lookup_table, 
        pl_lookup_table, pl_lookup_table_id, pl_lookup_table_name)

    return sql

def wh_get_base_df(base):
    sql = gen_sql_wh_percentage_overlap(base)
    return pd.read_sql(sql, engine)

def wh():
    wb = Workbook()

    # each base has a sheet
    for base in BASE_LOOKUP:
        if base not in BASE_LOOKUP:
            raise Exception('base not in the base_lookup')

        base_table = BASE_LOOKUP[base][0]
        base_id = BASE_LOOKUP[base][1]

        ws = wb.create_sheet(title='_'.join([base_table, base_id]))

        base_df = wh_get_base_df(base)
        for row in dataframe_to_rows(base_df, index=False):
            ws.append(row)

    wb.save('wh_2016_full.xlsx')


def tls():
    wb = Workbook()

    # each base has a sheet
    for base in BASE_LOOKUP:
        if base not in BASE_LOOKUP:
            raise Exception('base not in the base_lookup')

        base_table = BASE_LOOKUP[base][0]
        base_id = BASE_LOOKUP[base][1]

        ws = wb.create_sheet(title='_'.join([base_table, base_id]))

        base_df = tls_get_base_df(base)
        for row in dataframe_to_rows(base_df, index=False):
            ws.append(row)

    wb.save('tls_2016_full.xlsx')

COMBINED_WH_NOMINATION_VIEW = 'z_combined_wh_nomination_view'
TLS_SHAPE = 'tls.tentative'

OUTPUT_SCHEMA = 'ca_2016'
TLS_SCHEMA = 'ca_tls'
